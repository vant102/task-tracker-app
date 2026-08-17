"""
WorkLife Local Server (Port 8081)
Cung cấp API kết nối giữa Web App và máy tính cá nhân:
1. Mở thư mục dự án trên Windows Explorer (os.startfile)
2. Đồng bộ 2 chiều (Cloud Firestore <-> H:\\My Drive\\Worklife_NVT\\Worklife_Sync.xlsx)
3. Đọc dữ liệu từ file Excel trên máy tính
"""

import os
import sys
import json
import urllib.parse
from http.server import HTTPServer, BaseHTTPRequestHandler
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PORT = 8081
EXCEL_PATH = r"H:\My Drive\Worklife_NVT\Worklife_Sync.xlsx"

STANDARDIZED_COLUMNS = [
    "ID",
    "Tên Dự Án",
    "Trạng Thái",
    "Phạm Vi",
    "Loại Hình",
    "Bắt Đầu",
    "Kết Thúc",
    "Chốt Giai Đoạn",
    "Chủ Đầu Tư",
    "Địa Điểm",
    "Phong Cách",
    "Path",
    "IsOnline"
]

def save_projects_to_excel(projects_list):
    """Ghi danh sách dự án từ Web App vào file Excel Worklife_Sync.xlsx"""
    rows = []
    for p in projects_list:
        scopes = p.get('scope', [])
        scope_str = ', '.join(scopes) if isinstance(scopes, list) else str(scopes or '')
        
        status = p.get('status', 'Working')
        if p.get('completed') is True:
            status = 'Completed'
            
        rows.append({
            "ID": p.get('project_id_code', p.get('ID', p.get('id', ''))),
            "Tên Dự Án": p.get('name', p.get('Tên Dự Án', '')),
            "Trạng Thái": status,
            "Phạm Vi": scope_str,
            "Loại Hình": p.get('project_type', p.get('Loại Hình', '')),
            "Bắt Đầu": p.get('start_month', p.get('Bắt Đầu', '')),
            "Kết Thúc": p.get('end_month', p.get('Kết Thúc', '')),
            "Chốt Giai Đoạn": p.get('phase_deadline', p.get('Chốt Giai Đoạn', '')),
            "Chủ Đầu Tư": p.get('client', p.get('Chủ Đầu Tư', '')),
            "Địa Điểm": p.get('location', p.get('Địa Điểm', '')),
            "Phong Cách": p.get('style', p.get('Phong Cách', '')),
            "Path": p.get('local_path', p.get('Path', '')),
            "IsOnline": True
        })
        
    df = pd.DataFrame(rows, columns=STANDARDIZED_COLUMNS)
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    
    with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='All_Projects', index=False)
        
    # Format Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['All_Projects']
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="E6B965")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_num in range(1, len(STANDARDIZED_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(STANDARDIZED_COLUMNS)):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            if cell.column in [1, 3, 4, 5, 6, 7, 8, 13]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    col_widths = {'A': 14, 'B': 32, 'C': 16, 'D': 22, 'E': 20, 'F': 14, 'G': 14, 'H': 20, 'I': 22, 'J': 24, 'K': 20, 'L': 45, 'M': 12}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 28
    wb.save(EXCEL_PATH)
    return len(rows)

def read_projects_from_excel():
    """Đọc danh sách dự án từ file Excel"""
    if not os.path.exists(EXCEL_PATH):
        return []
    df = pd.read_excel(EXCEL_PATH, sheet_name='All_Projects')
    df = df.where(pd.notnull(df), "")
    return df.to_dict(orient='records')

class RequestHandler(BaseHTTPRequestHandler):
    def _send_cors_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, X-Requested-With')

    def do_OPTIONS(self):
        self.send_response(200)
        self._send_cors_headers()
        self.end_headers()

    def do_GET(self):
        parsed_url = urllib.parse.urlparse(self.path)
        params = urllib.parse.parse_qs(parsed_url.query)

        # 1. API: Mở thư mục trên Windows Explorer
        if parsed_url.path == '/open-folder':
            path = params.get('path', [''])[0]
            if path and os.path.exists(path):
                try:
                    os.startfile(path)
                    self.send_response(200)
                    self._send_cors_headers()
                    self.send_header('Content-Type', 'application/json; charset=utf-8')
                    self.end_headers()
                    self.wfile.write(json.dumps({'status': 'success', 'message': f'Opened {path}'}).encode('utf-8'))
                    return
                except Exception as e:
                    self.send_response(500)
                    self._send_cors_headers()
                    self.end_headers()
                    self.wfile.write(json.dumps({'status': 'error', 'message': str(e)}).encode('utf-8'))
                    return
            else:
                self.send_response(404)
                self._send_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'status': 'error', 'message': 'Path not found on PC'}).encode('utf-8'))
                return

        # 2. API: Đọc dữ liệu từ file Excel
        elif parsed_url.path == '/read-excel':
            try:
                projects = read_projects_from_excel()
                self.send_response(200)
                self._send_cors_headers()
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({'status': 'success', 'data': projects}, ensure_ascii=False).encode('utf-8'))
                return
            except Exception as e:
                self.send_response(500)
                self._send_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'status': 'error', 'message': str(e)}).encode('utf-8'))
                return

        # Default 404
        self.send_response(404)
        self._send_cors_headers()
        self.end_headers()

    def do_POST(self):
        parsed_url = urllib.parse.urlparse(self.path)

        # 3. API: Đồng bộ 2 chiều (Lưu dữ liệu Web App -> File Excel và trả về dữ liệu mới)
        if parsed_url.path == '/sync-excel':
            try:
                content_length = int(self.headers.get('Content-Length', 0))
                body = self.rfile.read(content_length)
                payload = json.loads(body.decode('utf-8'))
                
                web_projects = payload.get('projects', [])
                
                # Merge with existing Excel if needed or overwrite with clean records
                count = save_projects_to_excel(web_projects)
                
                self.send_response(200)
                self._send_cors_headers()
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'status': 'success', 
                    'message': f'Đã đồng bộ {count} dự án vào {EXCEL_PATH}',
                    'count': count
                }, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self._send_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({'status': 'error', 'message': str(e)}).encode('utf-8'))

def run_server():
    server_address = ('', PORT)
    httpd = HTTPServer(server_address, RequestHandler)
    print(f"=====================================================")
    print(f"🚀 WorkLife Local Server is running on port {PORT}")
    print(f"📂 Excel Target: {EXCEL_PATH}")
    print(f"=====================================================")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping server...")
        httpd.server_close()

if __name__ == '__main__':
    run_server()
